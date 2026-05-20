"""
Reports 탭 - 샘플별 QC 리포트 생성 및 내보내기
PDF : matplotlib PdfPages (추가 의존성 없음)
Excel: openpyxl
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import numpy as np
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QFont
from PyQt5.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTextBrowser,
    QVBoxLayout,
    QWidget,
)

try:
    import matplotlib
    matplotlib.use("Qt5Agg")
    import matplotlib.pyplot as plt
    import matplotlib.gridspec as gridspec
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

try:
    import openpyxl
    from openpyxl.styles import Font as XlFont, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from config.settings import QC_STEPS, RNA_QC_STEPS, STATUS_COLORS, REPORTS_DIR
from database import (
    db_manager,
    get_all_samples,
    get_qc_metrics_by_sample,
    get_smear_analyses_by_sample,
)

logger = logging.getLogger(__name__)

_STATUS_HEX = STATUS_COLORS          # {"Pass": "#4CAF50", ...}
_STATUS_MPL = {                       # matplotlib-safe colours
    "Pass":    "#4CAF50",
    "Warning": "#FF9800",
    "Fail":    "#F44336",
    "No Data": "#9E9E9E",
}

# ── 선택 목록 컬럼 ──────────────────────────────────────────────────
_SEL_COLS = ["", "Sample ID", "Name", "Type", "Latest Status"]


def _fmt(val, decimals: int = 2) -> str:
    if val is None:
        return "-"
    return f"{val:.{decimals}f}"


def _smear_low_high(step_smears: dict):
    """Parse RNA smear ranges into (pct_low, pct_high).

    Widest-span range = Total → excluded to avoid double-counting.
    Remaining: midpoint < 1000 → low, midpoint >= 1000 → high.
    Returns (pct_low, pct_high) floats, or (None, None) if no data.
    """
    import re

    def _span(rt):
        nums = [float(n) for n in re.findall(r'\d+(?:\.\d+)?',
                                              str(rt).replace(',', ''))]
        return (nums[1] - nums[0]) if len(nums) >= 2 else 0.0

    if not step_smears:
        return None, None

    total_key = max(step_smears, key=_span)
    pct_low = 0.0
    pct_high = 0.0
    has_data = False

    for rng_text, sa in step_smears.items():
        if rng_text == total_key:
            continue
        if sa.pct_total is None:
            continue
        text = str(rng_text).replace(',', '')
        if re.search(r'marker|ladder', text, re.IGNORECASE):
            continue
        nums = [float(n) for n in re.findall(r'\d+(?:\.\d+)?', text)]
        if not nums:
            continue
        start = nums[0]
        end = nums[1] if len(nums) >= 2 else start * 5
        mid = (start + end) / 2
        has_data = True
        if mid < 1000:
            pct_low += sa.pct_total
        else:
            pct_high += sa.pct_total

    return (pct_low, pct_high) if has_data else (None, None)


def _widest_cv(step_smears: dict) -> str:
    """Return %CV string from the widest-span smear range (= Total range)."""
    import re

    def _span(rt):
        nums = [float(n) for n in re.findall(r'\d+(?:\.\d+)?',
                                              str(rt).replace(',', ''))]
        return (nums[1] - nums[0]) if len(nums) >= 2 else 0.0

    if not step_smears:
        return "-"
    widest_key = max(step_smears, key=_span)
    sa = step_smears.get(widest_key)
    if sa is None or sa.cv is None:
        return "-"
    return f"{sa.cv:.1f}"


def _compute_mqi(step_smears: dict) -> str:
    """MQI = High% / (Low% + High%) — 0-1 scale, higher = more intact RNA."""
    pct_low, pct_high = _smear_low_high(step_smears)
    if pct_high is not None and pct_low is not None:
        denom = pct_low + pct_high
        if denom > 0:
            return f"{pct_high / denom:.2f}"
    return "-"


# ════════════════════════════════════════════════════════════════════
# Reports 탭 메인 위젯
# ════════════════════════════════════════════════════════════════════

class ReportsTab(QWidget):
    """Reports 탭"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._samples_cache: list = []          # DB Sample objects (session 밖 snapshot)
        self._preview_sample_id: Optional[str] = None
        self._build_ui()
        self.refresh()

    # ── UI 구성 ──────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        # 헤더 바
        hdr = QHBoxLayout()
        title = QLabel("Reports")
        f = QFont(); f.setPointSize(14); f.setBold(True)
        title.setFont(f)
        hdr.addWidget(title)
        hdr.addStretch()

        btn_refresh = QPushButton("Refresh")
        btn_refresh.clicked.connect(self.refresh)
        hdr.addWidget(btn_refresh)

        self.btn_pdf = QPushButton("Export PDF (selected)")
        self.btn_pdf.clicked.connect(self._export_pdf)
        self.btn_pdf.setEnabled(HAS_MPL)
        hdr.addWidget(self.btn_pdf)

        self.btn_excel = QPushButton("Export Excel (selected)")
        self.btn_excel.clicked.connect(self._export_excel)
        self.btn_excel.setEnabled(HAS_OPENPYXL)
        hdr.addWidget(self.btn_excel)

        self.btn_html = QPushButton("Export HTML (selected)")
        self.btn_html.clicked.connect(self._export_html)
        hdr.addWidget(self.btn_html)

        root.addLayout(hdr)

        # 좌우 스플리터
        splitter = QSplitter(Qt.Horizontal)
        self.splitter = splitter

        # ── 왼쪽: 샘플 선택 목록 ─────────────────────────────────
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(4)

        # 검색 + 프로젝트 + 타입 필터
        filter_row = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Search Sample ID…")
        self.search_edit.textChanged.connect(self._apply_filter)
        filter_row.addWidget(self.search_edit)

        self.proj_filter = QComboBox()
        self.proj_filter.addItem("All Projects")
        self.proj_filter.currentIndexChanged.connect(self._apply_filter)
        filter_row.addWidget(self.proj_filter)

        self.type_filter = QComboBox()
        self.type_filter.addItem("All Types")
        self.type_filter.currentIndexChanged.connect(self._apply_filter)
        filter_row.addWidget(self.type_filter)
        left_layout.addLayout(filter_row)

        # 전체선택 / 해제
        sel_row = QHBoxLayout()
        btn_all = QPushButton("Select All")
        btn_all.clicked.connect(self._select_all)
        sel_row.addWidget(btn_all)
        btn_none = QPushButton("Deselect All")
        btn_none.clicked.connect(self._deselect_all)
        sel_row.addWidget(btn_none)
        left_layout.addLayout(sel_row)

        # 샘플 선택 테이블
        self.sel_table = QTableWidget()
        self.sel_table.setColumnCount(len(_SEL_COLS))
        self.sel_table.setHorizontalHeaderLabels(_SEL_COLS)
        self.sel_table.setColumnWidth(0, 28)        # checkbox col
        self.sel_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.sel_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Interactive)
        self.sel_table.horizontalHeader().setStretchLastSection(True)
        self.sel_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.sel_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.sel_table.setAlternatingRowColors(True)
        self.sel_table.setSortingEnabled(True)
        self.sel_table.horizontalHeader().setSortIndicatorShown(True)
        self.sel_table.verticalHeader().setVisible(False)
        self.sel_table.currentCellChanged.connect(self._on_row_selected)
        self.sel_table.itemChanged.connect(self._on_checkbox_changed)
        left_layout.addWidget(self.sel_table)

        splitter.addWidget(left)

        # ── 오른쪽: 미리보기 ─────────────────────────────────────
        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(4)

        preview_lbl = QLabel("Preview")
        f2 = QFont(); f2.setBold(True)
        preview_lbl.setFont(f2)
        right_layout.addWidget(preview_lbl)

        # 정보 패널 (HTML)
        self.info_browser = QTextBrowser()
        self.info_browser.setMaximumHeight(160)
        self.info_browser.setOpenExternalLinks(False)
        right_layout.addWidget(self.info_browser)

        # matplotlib 차트 (미리보기)
        if HAS_MPL:
            self._prev_fig, self._prev_axes = plt.subplots(
                1, 2, figsize=(8, 3.2), dpi=88
            )
            self._prev_fig.patch.set_facecolor("#FAFAFA")
            self._prev_canvas = FigureCanvas(self._prev_fig)
            self._prev_canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            right_layout.addWidget(self._prev_canvas, 1)
        else:
            right_layout.addWidget(QLabel("matplotlib unavailable – chart disabled."), 1)

        # QC 상세 테이블 (미리보기)
        self.qc_preview = QTableWidget()
        qc_cols = ["Step", "Instrument", "Conc (ng/ul)", "Vol (ul)",
                   "Total (ng)", "GQN/RIN", "Avg Size (bp)", "Status", "Date"]
        self.qc_preview.setColumnCount(len(qc_cols))
        self.qc_preview.setHorizontalHeaderLabels(qc_cols)
        self.qc_preview.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.qc_preview.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.qc_preview.horizontalHeader().setStretchLastSection(True)
        self.qc_preview.setMaximumHeight(180)
        right_layout.addWidget(self.qc_preview)

        splitter.addWidget(right)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)

        root.addWidget(splitter, 1)

    # ── 데이터 로드 ──────────────────────────────────────────────────

    def refresh(self):
        """DB에서 샘플 목록 새로고침."""
        try:
            with db_manager.session_scope() as session:
                samples = get_all_samples(session)

                # session 밖에서 쓸 수 있도록 dict 형태로 스냅샷
                # get_all_samples은 created_at DESC 정렬 → 첫 등장이 최신 레코드
                _seen_ids: set = set()
                _cache_raw = []
                for s in samples:
                    if s.sample_id in _seen_ids:
                        logger.warning(f"Duplicate sample_id in DB: {s.sample_id} — skipping older record")
                        continue
                    _seen_ids.add(s.sample_id)
                    _cache_raw.append({
                        "sample_id": s.sample_id,
                        "sample_name": s.sample_name or "",
                        "sample_type": s.sample_type or "",
                        "species": getattr(s, "species", None) or "",
                        "material": getattr(s, "material", None) or "",
                        "description": s.description or "",
                        "project": getattr(s, "project", None) or "",
                    })
                self._samples_cache = _cache_raw

                # 전체 QC 레코드 중 가장 나쁜 status를 overall로 사용
                _sorder = {"Fail": 0, "Warning": 1, "Pass": 2}
                for snap in self._samples_cache:
                    all_m = get_qc_metrics_by_sample(session, snap["sample_id"])
                    worst = None
                    for m in all_m:
                        if m.status in _sorder:
                            if worst is None or _sorder[m.status] < _sorder[worst]:
                                worst = m.status
                    snap["latest_status"] = worst or "No Data"

        except Exception as e:
            logger.error(f"Reports refresh failed: {e}")
            self._samples_cache = []

        # 프로젝트 필터 콤보 갱신
        projects = sorted({s["project"] for s in self._samples_cache if s["project"]})
        self.proj_filter.blockSignals(True)
        current_proj = self.proj_filter.currentText()
        self.proj_filter.clear()
        self.proj_filter.addItem("All Projects")
        for p in projects:
            self.proj_filter.addItem(p)
        idx = self.proj_filter.findText(current_proj)
        if idx >= 0:
            self.proj_filter.setCurrentIndex(idx)
        self.proj_filter.blockSignals(False)

        # 타입 필터 콤보 갱신
        types = sorted({s["sample_type"] for s in self._samples_cache if s["sample_type"]})
        self.type_filter.blockSignals(True)
        current_type = self.type_filter.currentText()
        self.type_filter.clear()
        self.type_filter.addItem("All Types")
        for t in types:
            self.type_filter.addItem(t)
        idx = self.type_filter.findText(current_type)
        if idx >= 0:
            self.type_filter.setCurrentIndex(idx)
        self.type_filter.blockSignals(False)

        self._apply_filter()

    def _apply_filter(self):
        """검색어 + 프로젝트 + 타입 필터 적용."""
        keyword  = self.search_edit.text().strip().lower()
        proj_sel = self.proj_filter.currentText()
        type_sel = self.type_filter.currentText()

        filtered = [
            s for s in self._samples_cache
            if (not keyword or keyword in s["sample_id"].lower()
                            or keyword in s["sample_name"].lower())
            and (proj_sel == "All Projects" or s["project"] == proj_sel)
            and (type_sel == "All Types"    or s["sample_type"] == type_sel)
        ]

        # 기존 checked 상태 보존
        checked = self._get_checked_ids()

        self.sel_table.blockSignals(True)
        self.sel_table.setRowCount(len(filtered))
        for row, s in enumerate(filtered):
            # checkbox
            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            chk.setCheckState(Qt.Checked if s["sample_id"] in checked else Qt.Unchecked)
            self.sel_table.setItem(row, 0, chk)

            self.sel_table.setItem(row, 1, QTableWidgetItem(s["sample_id"]))
            self.sel_table.setItem(row, 2, QTableWidgetItem(s["sample_name"]))
            self.sel_table.setItem(row, 3, QTableWidgetItem(s["sample_type"]))

            status = s["latest_status"]
            st_item = QTableWidgetItem(status)
            color = _STATUS_HEX.get(status)
            if color:
                st_item.setForeground(QColor(color))
            self.sel_table.setItem(row, 4, st_item)

        self.sel_table.blockSignals(False)

    def _get_checked_ids(self) -> list:
        """테이블 표시 순서를 유지한 채 선택된 Sample ID 목록 반환."""
        ids = []
        for row in range(self.sel_table.rowCount()):
            chk = self.sel_table.item(row, 0)
            sid = self.sel_table.item(row, 1)
            if chk and sid and chk.checkState() == Qt.Checked:
                ids.append(sid.text())
        return ids

    def _select_all(self):
        self.sel_table.blockSignals(True)
        for row in range(self.sel_table.rowCount()):
            chk = self.sel_table.item(row, 0)
            if chk:
                chk.setCheckState(Qt.Checked)
        self.sel_table.blockSignals(False)

    def _deselect_all(self):
        self.sel_table.blockSignals(True)
        for row in range(self.sel_table.rowCount()):
            chk = self.sel_table.item(row, 0)
            if chk:
                chk.setCheckState(Qt.Unchecked)
        self.sel_table.blockSignals(False)

    def _on_checkbox_changed(self, item):
        pass  # 나중에 필요하면 확장

    def _on_row_selected(self, row, _col, _prev_row, _prev_col):
        if row < 0:
            return
        sid_item = self.sel_table.item(row, 1)
        if sid_item:
            self._load_preview(sid_item.text())

    # ── 미리보기 ─────────────────────────────────────────────────────

    def _load_preview(self, sample_id: str):
        self._preview_sample_id = sample_id

        snap = next((s for s in self._samples_cache if s["sample_id"] == sample_id), None)
        if not snap:
            return

        try:
            with db_manager.session_scope() as session:
                metrics = get_qc_metrics_by_sample(session, sample_id)
                metrics_dicts = [
                    {
                        "step": m.step,
                        "instrument": m.instrument,
                        "concentration": m.concentration,
                        "volume": m.volume,
                        "total_amount": m.total_amount,
                        "gqn_rin": m.gqn_rin,
                        "avg_size": m.avg_size,
                        "status": m.status,
                        "measured_at": m.measured_at,
                    }
                    for m in metrics
                ]
        except Exception as e:
            logger.error(f"Preview load failed: {e}")
            metrics_dicts = []

        self._render_info_html(snap, metrics_dicts)
        self._render_qc_table(metrics_dicts)
        if HAS_MPL:
            self._render_preview_chart(sample_id, metrics_dicts)

    def _render_info_html(self, snap: dict, metrics: list):
        status = snap["latest_status"]
        color = _STATUS_HEX.get(status, "#9E9E9E")
        html = f"""
        <table style="font-size:11px; border-collapse:collapse; width:100%">
          <tr><td><b>Sample ID</b></td><td>{snap['sample_id']}</td>
              <td><b>Name</b></td><td>{snap['sample_name'] or '-'}</td></tr>
          <tr><td><b>Type</b></td><td>{snap['sample_type']}</td>
              <td><b>Latest Status</b></td>
              <td><span style="color:{color}; font-weight:bold">{status}</span></td></tr>
          <tr><td><b>Species</b></td><td>{snap['species'] or '-'}</td>
              <td><b>Material</b></td><td>{snap['material'] or '-'}</td></tr>
          <tr><td><b>Description</b></td>
              <td colspan="3">{snap['description'] or '-'}</td></tr>
          <tr><td><b>QC Records</b></td><td>{len(metrics)}</td>
              <td><b>Steps covered</b></td>
              <td>{', '.join(sorted({m['step'] for m in metrics if m['step']})) or '-'}</td></tr>
        </table>
        """
        self.info_browser.setHtml(html)

    def _render_qc_table(self, metrics: list):
        self.qc_preview.setRowCount(len(metrics))
        for row, m in enumerate(metrics):
            self.qc_preview.setItem(row, 0, QTableWidgetItem(m["step"] or "-"))
            self.qc_preview.setItem(row, 1, QTableWidgetItem(m["instrument"] or "-"))
            self.qc_preview.setItem(row, 2, QTableWidgetItem(_fmt(m["concentration"])))
            self.qc_preview.setItem(row, 3, QTableWidgetItem(_fmt(m["volume"])))
            self.qc_preview.setItem(row, 4, QTableWidgetItem(_fmt(m["total_amount"])))
            self.qc_preview.setItem(row, 5, QTableWidgetItem(_fmt(m["gqn_rin"])))
            self.qc_preview.setItem(row, 6, QTableWidgetItem(_fmt(m["avg_size"], 0)))

            status = m["status"] or "-"
            st_item = QTableWidgetItem(status)
            color = _STATUS_HEX.get(status)
            if color:
                st_item.setForeground(QColor(color))
            self.qc_preview.setItem(row, 7, st_item)

            date_str = (
                m["measured_at"].strftime("%Y-%m-%d")
                if m["measured_at"] else "-"
            )
            self.qc_preview.setItem(row, 8, QTableWidgetItem(date_str))

    def _render_preview_chart(self, sample_id: str, metrics: list):
        ax1, ax2 = self._prev_axes
        ax1.clear(); ax2.clear()

        if not metrics:
            for ax in (ax1, ax2):
                ax.text(0.5, 0.5, "No QC data", ha="center", va="center",
                        transform=ax.transAxes, color="#9E9E9E")
                ax.axis("off")
            self._prev_fig.tight_layout(pad=0.5)
            self._prev_canvas.draw()
            return

        # Two-line x-tick labels: step name + instrument abbreviation
        steps = [
            f"{(m['step'] or '?')}\n({_short_instrument(m.get('instrument', ''))})"
            if m.get('instrument') else (m['step'] or '?')
            for m in metrics
        ]
        concs = [m["concentration"] or 0 for m in metrics]
        gqns  = [m["gqn_rin"] or 0 for m in metrics]
        statuses = [m["status"] or "No Data" for m in metrics]
        bar_colors = [_STATUS_MPL.get(s, "#9E9E9E") for s in statuses]
        x = np.arange(len(steps))

        # 왼쪽: 농도
        bars1 = ax1.bar(x, concs, color=bar_colors, edgecolor="white", linewidth=0.8)
        ax1.set_title("Concentration (ng/µl)", fontsize=9, fontweight="bold")
        ax1.set_xticks(x)
        ax1.set_xticklabels(steps, rotation=35, ha="right", fontsize=7, linespacing=1.2)
        ax1.tick_params(axis="y", labelsize=7)
        ax1.spines[["top", "right"]].set_visible(False)
        ax1.set_facecolor("#FAFAFA")
        max_c = max((v for v in concs if v), default=0)
        if max_c > 0:
            ax1.set_ylim(0, max_c * 1.28)
        for bar, val in zip(bars1, concs):
            if val > 0:
                ax1.text(bar.get_x() + bar.get_width() / 2,
                         bar.get_height(), f"{val:.1f}",
                         ha="center", va="bottom", fontsize=7)

        # 오른쪽: GQN/RIN
        bars2 = ax2.bar(x, gqns, color=bar_colors, edgecolor="white", linewidth=0.8)
        ax2.set_title("GQN / RIN", fontsize=9, fontweight="bold")
        ax2.set_xticks(x)
        ax2.set_xticklabels(steps, rotation=35, ha="right", fontsize=7, linespacing=1.2)
        ax2.tick_params(axis="y", labelsize=7)
        ax2.spines[["top", "right"]].set_visible(False)
        ax2.set_facecolor("#FAFAFA")
        max_g = max((v for v in gqns if v), default=0)
        if max_g > 0:
            ax2.set_ylim(0, max_g * 1.28)
        for bar, val in zip(bars2, gqns):
            if val > 0:
                ax2.text(bar.get_x() + bar.get_width() / 2,
                         bar.get_height(), f"{val:.1f}",
                         ha="center", va="bottom", fontsize=7)

        self._prev_fig.suptitle(f"QC Preview — {sample_id}", fontsize=9, y=1.01)
        self._prev_fig.tight_layout(pad=1.2)
        self._prev_canvas.draw()

    # ── PDF 내보내기 ─────────────────────────────────────────────────

    def _export_pdf(self):
        selected_ids = self._get_checked_ids()   # 테이블 순서 유지
        if not selected_ids:
            QMessageBox.information(self, "No Selection",
                                    "Check at least one sample in the list.")
            return

        default_name = f"QC_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save PDF Report",
            str(REPORTS_DIR / default_name),
            "PDF Files (*.pdf)",
        )
        if not save_path:
            return

        snap_map = {s["sample_id"]: s for s in self._samples_cache}
        failed = []

        try:
            with PdfPages(save_path) as pdf:
                # bbox_inches 없이 저장해야 모든 페이지가 정확히 A4(8.27×11.69)로 통일됨

                # 1) 표지
                cover_fig = _build_cover_page(selected_ids, snap_map)
                pdf.savefig(cover_fig)
                plt.close(cover_fig)

                # 2) 배치 요약 (step-status table + 3 batch charts)
                overview_fig = _build_batch_overview(selected_ids, snap_map)
                pdf.savefig(overview_fig)
                plt.close(overview_fig)

                # 3) 샘플별 페이지 (차트 3개 + QC 표 + 전기영동 한 페이지)
                for sid in selected_ids:
                    snap = snap_map.get(sid, {
                        "sample_id": sid, "sample_name": "",
                        "sample_type": "", "species": "",
                        "material": "", "description": "",
                        "latest_status": "",
                    })
                    try:
                        with db_manager.session_scope() as session:
                            metrics = get_qc_metrics_by_sample(session, sid)
                            metrics_dicts = [
                                {
                                    "step": m.step,
                                    "instrument": m.instrument,
                                    "concentration": m.concentration,
                                    "volume": m.volume,
                                    "total_amount": m.total_amount,
                                    "gqn_rin": m.gqn_rin,
                                    "avg_size": m.avg_size,
                                    "peak_size": m.peak_size,
                                    "status": m.status,
                                    "measured_at": m.measured_at,
                                }
                                for m in metrics
                            ]
                            fig = _build_sample_combined_page(sid, snap, metrics_dicts, session)
                            pdf.savefig(fig)
                            plt.close(fig)

                    except Exception as e:
                        logger.error(f"PDF page failed for {sid}: {e}")
                        failed.append(sid)

            msg = f"PDF saved:\n{save_path}\n\n{len(selected_ids) - len(failed)} sample(s) exported."
            if failed:
                msg += f"\nFailed: {', '.join(failed)}"
            QMessageBox.information(self, "Export Complete", msg)

        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to save PDF:\n{e}")

    # ── Excel 내보내기 ────────────────────────────────────────────────

    def _export_excel(self):
        selected_ids = self._get_checked_ids()   # 테이블 순서 유지
        if not selected_ids:
            QMessageBox.information(self, "No Selection",
                                    "Check at least one sample in the list.")
            return

        default_name = f"QC_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File",
            str(REPORTS_DIR / default_name),
            "Excel Files (*.xlsx)",
        )
        if not save_path:
            return

        snap_map = {s["sample_id"]: s for s in self._samples_cache}

        try:
            wb = openpyxl.Workbook()
            _write_excel(wb, selected_ids, snap_map)
            wb.save(save_path)
            QMessageBox.information(self, "Export Complete",
                                    f"Excel saved:\n{save_path}")
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to save Excel:\n{e}")

    # ── HTML 내보내기 ─────────────────────────────────────────────────

    def _export_html(self):
        selected_ids = self._get_checked_ids()
        if not selected_ids:
            QMessageBox.information(self, "No Selection",
                                    "Check at least one sample in the list.")
            return

        folder = QFileDialog.getExistingDirectory(
            self, "Select Output Folder", str(REPORTS_DIR)
        )
        if not folder:
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_dir = Path(folder) / f"NGS_QC_Report_{timestamp}"
        report_dir.mkdir(parents=True, exist_ok=True)

        snap_map = {s["sample_id"]: s for s in self._samples_cache}

        try:
            from ui.html_report import generate_html_report
            generate_html_report(selected_ids, snap_map, report_dir)
            index_path = report_dir / "index.html"
            QMessageBox.information(
                self, "Export Complete",
                f"HTML report saved to:\n{index_path}"
            )
        except Exception as e:
            logger.error(f"HTML export failed: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to save HTML:\n{e}")

    # ── GUI 상태 저장/복원 ────────────────────────────────────────────

    def save_gui_state(self, settings):
        from config.gui_state import save_table_widths, save_splitter, save_combo
        save_table_widths(settings, "ReportsTab/selTableWidths",     self.sel_table)
        save_table_widths(settings, "ReportsTab/qcPreviewWidths",    self.qc_preview)
        save_splitter(settings,     "ReportsTab/splitterState",      self.splitter)
        save_combo(settings,        "ReportsTab/projFilter",         self.proj_filter)
        save_combo(settings,        "ReportsTab/typeFilter",         self.type_filter)
        settings.setValue("ReportsTab/searchText", self.search_edit.text())

    def restore_gui_state(self, settings):
        from config.gui_state import restore_table_widths, restore_splitter, restore_combo
        restore_table_widths(settings, "ReportsTab/selTableWidths",  self.sel_table)
        restore_table_widths(settings, "ReportsTab/qcPreviewWidths", self.qc_preview)
        restore_splitter(settings,     "ReportsTab/splitterState",   self.splitter)
        self.proj_filter.blockSignals(True)
        restore_combo(settings,        "ReportsTab/projFilter",      self.proj_filter)
        self.proj_filter.blockSignals(False)
        self.type_filter.blockSignals(True)
        restore_combo(settings,        "ReportsTab/typeFilter",      self.type_filter)
        self.type_filter.blockSignals(False)
        text = settings.value("ReportsTab/searchText", "")
        if text:
            self.search_edit.setText(str(text))


# ════════════════════════════════════════════════════════════════════
# PDF 빌더 함수 (모듈 레벨)
# ════════════════════════════════════════════════════════════════════

def _criteria_note(sample_type: str) -> str:
    """샘플 타입별 QC 판정 기준 한 줄 주석 — config/settings.py에서 동적 로드."""
    from config.settings import QC_CRITERIA
    if sample_type == "WGS":
        gqn_c = QC_CRITERIA.get("WGS", {}).get("GQN", {})
        gp = gqn_c.get("pass", 7.0)
        gw = gqn_c.get("warning", 5.0)
        return (
            f"판정 기준 ┃ GQN (Femto Pulse) : Pass ≥ {gp}  /  Warning {gw} – {gp - 0.1:.1f}  /  Fail < {gw}"
            "   ┃  농도 (Qubit/NanoDrop) : 참고용 (판정 미사용)"
        )
    elif sample_type == "mRNA-seq":
        from config.settings import QC_CRITERIA
        rin_c = QC_CRITERIA.get("mRNA-seq", {}).get("RIN", {})
        rp = rin_c.get("pass", 7.0)
        rw = rin_c.get("warning", 5.0)
        return (
            f"판정 기준 ┃ RIN (Femto Pulse) : Pass ≥ {rp}  /  Warning {rw} – {rp - 0.1:.1f}  /  Fail < {rw}"
            "   ┃  Total RNA (Qubit/NanoDrop) : Pass ≥ 1,000 ng (1 µg)  /  Warning < 1,000 ng  /  Fail 없음"
        )
    return ""


def _instrument_rank(instrument: Optional[str]) -> int:
    """Qubit(0) > NanoDrop(1) > FemtoPulse(2) > 기타(99)."""
    if not instrument:
        return 99
    i = instrument.lower()
    if "qubit" in i:
        return 0
    if "nanodrop" in i or "nano" in i:
        return 1
    if "femto" in i:
        return 2
    return 3


def _short_instrument(instrument: Optional[str]) -> str:
    """Abbreviated instrument label for bar chart x-axis."""
    if not instrument:
        return ""
    i = instrument.lower()
    if "qubit" in i:
        return "Qubit"
    if "nanodrop" in i or "nano" in i:
        return "NanoDrop"
    if "femto" in i:
        return "FP"
    return instrument[:8]


def _collect_per_sample_data(sample_ids: List[str]) -> dict:
    """샘플별 total_amount / gqn_rin / concentration 수집.

    concentration은 Qubit > NanoDrop > FemtoPulse 우선순위로 선택.
    반환: {sid: {total_amount, amount_status, gqn_rin, rin_status,
                 concentration, conc_status}}
    """
    result: dict = {}
    for sid in sample_ids:
        data: dict = {
            "total_amount": None, "amount_status": "No Data",
            "gqn_rin":      None, "rin_status":    "No Data",
            "concentration": None, "conc_status":  "No Data",
        }
        best_conc_rank = 99
        try:
            with db_manager.session_scope() as session:
                metrics = get_qc_metrics_by_sample(session, sid)
                # measured_at 기준 최신순 정렬
                sorted_metrics = sorted(
                    metrics,
                    key=lambda m: m.measured_at or datetime.min,
                    reverse=True,
                )
                for m in sorted_metrics:
                    if m.total_amount is not None and data["total_amount"] is None:
                        data["total_amount"] = m.total_amount
                        data["amount_status"] = m.status or "No Data"
                    if m.gqn_rin is not None and data["gqn_rin"] is None:
                        data["gqn_rin"] = m.gqn_rin
                        data["rin_status"] = m.status or "No Data"
                    rank = _instrument_rank(m.instrument)
                    if m.concentration is not None and rank < best_conc_rank:
                        data["concentration"] = m.concentration
                        best_conc_rank = rank
                        data["conc_status"] = m.status or "No Data"
        except Exception:
            pass
        result[sid] = data
    return result


def _bar_chart(ax, sample_ids: List[str], values: list, statuses: list,
               ylabel: str, title: str, fmt: str = "{:.1f}",
               thresholds: Optional[list] = None):
    """샘플별 bar chart 공통 헬퍼.

    thresholds: [(value, color, label), ...]
    """
    x = np.arange(len(sample_ids))
    colors = [_STATUS_MPL.get(s, "#9E9E9E") for s in statuses]
    bars = ax.bar(x, [v if v is not None else 0 for v in values],
                  color=colors, edgecolor="white", linewidth=0.4, alpha=0.88)

    if thresholds:
        for thr_val, thr_color, thr_label in thresholds:
            ax.axhline(thr_val, color=thr_color, linewidth=0.9,
                       linestyle="--", alpha=0.85, label=thr_label)
        ax.legend(fontsize=5.5, loc="upper right",
                  framealpha=0.7, edgecolor="none")

    # 값 레이블
    for bar, val in zip(bars, values):
        if val is not None and val > 0:
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height(),
                    fmt.format(val),
                    ha="center", va="bottom", fontsize=5.5, color="#333333")

    short_ids = [sid.split("-")[-1] if "-" in sid else sid for sid in sample_ids]
    ax.set_xticks(x)
    ax.set_xticklabels(short_ids, rotation=45, ha="right", fontsize=6)
    ax.set_ylabel(ylabel, fontsize=6.5)
    ax.tick_params(axis="y", labelsize=6)
    ax.spines[["top", "right"]].set_visible(False)
    ax.set_facecolor("#FAFAFA")
    ax.set_title(title, fontsize=8, fontweight="bold", loc="left", color="#1A237E")


def _build_batch_overview(sample_ids: List[str], snap_map: dict) -> "plt.Figure":
    """PDF 첫 페이지: step-status table + 3 per-sample charts. Portrait A4."""
    _order = {"Pass": 0, "Warning": 1, "Fail": 2}
    step_statuses: dict = {}

    for sid in sample_ids:
        by_step: dict = {}
        try:
            with db_manager.session_scope() as session:
                metrics = get_qc_metrics_by_sample(session, sid)
                for m in metrics:
                    s = m.status
                    if s in _order:
                        prev = by_step.get(m.step)
                        if prev is None or _order[s] > _order.get(prev, -1):
                            by_step[m.step] = s
        except Exception:
            pass
        step_statuses[sid] = by_step

    steps_with_data: set = set()
    for by_step in step_statuses.values():
        steps_with_data.update(by_step.keys())
    step_list = [s for s in _ALL_STEPS_ORDERED if s in steps_with_data]
    for s in steps_with_data:
        if s not in step_list:
            step_list.append(s)
    step_labels = [_STEP_ABBREV.get(s, s) for s in step_list]

    info_cols = ["Sample ID", "Name", "Type"]
    tbl_cols = info_cols + step_labels + ["Overall"]
    n_info = len(info_cols)

    tbl_data = []
    for sid in sample_ids:
        s = snap_map.get(sid, {})
        row = [sid, s.get("sample_name", "-"), s.get("sample_type", "-")]
        for step in step_list:
            row.append(step_statuses.get(sid, {}).get(step, "-"))
        row.append(s.get("latest_status", "-"))
        tbl_data.append(row)

    status_counts = {"Pass": 0, "Warning": 0, "Fail": 0, "No Data": 0}
    for sid in sample_ids:
        st = snap_map.get(sid, {}).get("latest_status", "No Data")
        status_counts[st] = status_counts.get(st, 0) + 1

    # ── per-sample 지표 수집 ─────────────────────────────────────────
    per_sample = _collect_per_sample_data(sample_ids)

    amounts  = [per_sample[sid]["total_amount"]  for sid in sample_ids]
    amt_st   = [per_sample[sid]["amount_status"] for sid in sample_ids]
    rins     = [per_sample[sid]["gqn_rin"]       for sid in sample_ids]
    rin_st   = [per_sample[sid]["rin_status"]    for sid in sample_ids]
    concs    = [per_sample[sid]["concentration"] for sid in sample_ids]
    conc_st  = [per_sample[sid]["conc_status"]   for sid in sample_ids]

    # ── Figure (portrait A4) ─────────────────────────────────────────
    fig = plt.figure(figsize=(8.27, 11.69))
    fig.patch.set_facecolor("white")

    gs = gridspec.GridSpec(
        5, 1, figure=fig,
        height_ratios=[0.4, 2.2, 1.3, 1.3, 1.3],
        hspace=0.65,
        left=0.07, right=0.95, top=0.963, bottom=0.034,
    )

    # ── 헤더 ─────────────────────────────────────────────────────────
    ax_hdr = fig.add_subplot(gs[0])
    ax_hdr.axis("off")
    ax_hdr.text(0, 1.0, "NGS Sample QC Report — Batch Overview",
                transform=ax_hdr.transAxes, fontsize=13, fontweight="bold",
                va="top", color="#1A237E")
    ax_hdr.text(0, 0.40,
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                f"Total: {len(sample_ids)} sample(s)",
                transform=ax_hdr.transAxes, fontsize=8, va="top", color="#555555")
    stat_x = 0.0
    for st in ("Pass", "Warning", "Fail", "No Data"):
        cnt = status_counts.get(st, 0)
        ax_hdr.text(stat_x, 0.0, f"{st}: {cnt}",
                    transform=ax_hdr.transAxes, fontsize=8, fontweight="bold",
                    color=_STATUS_MPL.get(st, "#9E9E9E"), va="bottom")
        stat_x += 0.22

    # ── Step-status table ─────────────────────────────────────────────
    ax_tbl = fig.add_subplot(gs[1])
    ax_tbl.axis("off")
    ax_tbl.set_title("QC Step Status", fontsize=9, fontweight="bold",
                     loc="left", pad=4, color="#1A237E")
    if tbl_data:
        tbl = ax_tbl.table(
            cellText=tbl_data, colLabels=tbl_cols,
            loc="upper center", cellLoc="center",
            bbox=[0.0, 0.0, 1.0, 0.96],
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(6.0)
        tbl.auto_set_column_width(list(range(len(tbl_cols))))
        for col in range(len(tbl_cols)):
            cell = tbl[0, col]
            cell.set_facecolor("#1A237E")
            cell.set_text_props(color="white", fontweight="bold", fontsize=6.0)
            cell.set_height(cell.get_height() * 1.5)
        for row_idx, row_data in enumerate(tbl_data):
            for col_idx in range(n_info, len(tbl_cols)):
                st = row_data[col_idx]
                color = _STATUS_MPL.get(st)
                if color and st != "-":
                    tbl[row_idx + 1, col_idx].set_facecolor(color)
                    tbl[row_idx + 1, col_idx].set_text_props(color="white", fontweight="bold")
            if row_idx % 2 == 1:
                for col in range(n_info):
                    tbl[row_idx + 1, col].set_facecolor("#F3F4F6")

    # ── Chart 1: Total Amount ─────────────────────────────────────────
    ax_amt = fig.add_subplot(gs[2])
    if any(v is not None for v in amounts):
        _bar_chart(ax_amt, sample_ids, amounts, amt_st,
                   "Total Amount (ng)", "Total Amount (ng)", fmt="{:.0f}")
    else:
        ax_amt.text(0.5, 0.5, "No total amount data", ha="center", va="center",
                    transform=ax_amt.transAxes, color="#9E9E9E", fontsize=8)
        ax_amt.axis("off")

    # ── Chart 2: GQN / RIN ───────────────────────────────────────────
    ax_rin = fig.add_subplot(gs[3])
    if any(v is not None for v in rins):
        from config.settings import QC_CRITERIA as _QC_C
        _rin_c = _QC_C.get("mRNA-seq", {}).get("RIN", {})
        _rp = _rin_c.get("pass", 7.0)
        _rw = _rin_c.get("warning", 5.0)
        _bar_chart(ax_rin, sample_ids, rins, rin_st,
                   "GQN / RIN", "GQN / RIN", fmt="{:.1f}",
                   thresholds=[
                       (_rp, _STATUS_MPL["Pass"],    f"Pass ≥{_rp}"),
                       (_rw, _STATUS_MPL["Warning"], f"Warning ≥{_rw}"),
                   ])
    else:
        ax_rin.text(0.5, 0.5, "No GQN/RIN data", ha="center", va="center",
                    transform=ax_rin.transAxes, color="#9E9E9E", fontsize=8)
        ax_rin.axis("off")

    # ── Chart 3: Concentration ───────────────────────────────────────
    ax_conc = fig.add_subplot(gs[4])
    if any(v is not None for v in concs):
        _bar_chart(ax_conc, sample_ids, concs, conc_st,
                   "Concentration (ng/µl)", "Concentration (ng/µl)  [Qubit > NanoDrop > FemtoPulse]",
                   fmt="{:.2f}")
    else:
        ax_conc.text(0.5, 0.5, "No concentration data", ha="center", va="center",
                     transform=ax_conc.transAxes, color="#9E9E9E", fontsize=8)
        ax_conc.axis("off")

    fig.text(0.5, 0.008, f"NGS Sample QC LIMS  |  {datetime.now().strftime('%Y-%m-%d')}",
             ha="center", fontsize=7, color="#888888")
    return fig


def _apply_table_style(tbl, n_cols: int, n_data_rows: int,
                       status_col: int = -1, metrics_dicts_ref: list = None,
                       font_size: float = 7.0,
                       hdr_color: str = "#2C3E6B",
                       alt_color: str = "#F4F6FB",
                       edge_color: str = "#D5D9E0"):
    """Clean, minimal table style: compact header, light alternating rows, thin borders."""
    ROW_H = 0.048   # data row height (axes fraction units — overridden by bbox)

    for col in range(n_cols):
        cell = tbl[0, col]
        cell.set_facecolor(hdr_color)
        cell.set_text_props(color="white", fontweight="bold", fontsize=font_size)
        cell.set_edgecolor(edge_color)
        cell.set_linewidth(0.4)

    for r in range(n_data_rows):
        row_bg = alt_color if r % 2 == 1 else "white"
        for col in range(n_cols):
            if (r + 1, col) not in tbl._cells:
                continue
            cell = tbl[r + 1, col]
            cell.set_facecolor(row_bg)
            cell.set_edgecolor(edge_color)
            cell.set_linewidth(0.4)
            cell.set_text_props(fontsize=font_size)

    # Status column: light tinted background + bold colored text
    if status_col >= 0 and metrics_dicts_ref is not None:
        _tint = {"Pass": "#E8F5E9", "Warning": "#FFF8E1", "Fail": "#FFEBEE"}
        _text = {"Pass": "#2E7D32", "Warning": "#E65100", "Fail": "#C62828"}
        for r, m in enumerate(metrics_dicts_ref):
            st = m.get("status", "") if isinstance(m, dict) else ""
            if st in _tint and (r + 1, status_col) in tbl._cells:
                tbl[r + 1, status_col].set_facecolor(_tint[st])
                tbl[r + 1, status_col].set_text_props(
                    color=_text[st], fontweight="bold", fontsize=font_size)


def _build_sample_combined_page(
    sid: str,
    snap: dict,
    metrics_dicts: list,
    session,
) -> "plt.Figure":
    """샘플 1개: QC Metrics 표 + Electropherogram + Smear Analysis 표. Portrait A4.

    Layout (y=0 bottom, y=1 top, A4 11.69 inch):
      header        0.950 ~ 0.982
      divider       0.944
      QC Metrics    label y=0.940 / table [0.048, 0.710, 0.904, 0.226]
      criteria note y=0.703 (va=top)
      divider       0.692
      Electropherogram label y=0.687 / axes [0.048, 0.352, 0.904, 0.330]
      Smear Analysis label y=0.344 / axes [0.048, 0.038, 0.904, 0.300]
      footer        y=0.016
    """
    import io

    traces, calibration = [], None
    try:
        from analysis.visualizer import load_electropherogram_traces, qc_visualizer
        traces, calibration = load_electropherogram_traces(sid)
    except Exception as e:
        logger.warning(f"Electropherogram load failed for {sid}: {e}")
    has_electro = bool(traces)

    smears = []
    try:
        smears = get_smear_analyses_by_sample(session, sid)
    except Exception:
        pass
    has_smear = bool(smears)

    status = snap.get("latest_status", "")
    status_color = _STATUS_MPL.get(status, "#9E9E9E")

    fig = plt.figure(figsize=(8.27, 11.69))
    fig.patch.set_facecolor("white")

    # ── 헤더 ──────────────────────────────────────────────────────────
    fig.text(0.048, 0.977,
             f"{snap.get('sample_id', '')}  —  {snap.get('sample_name', '')}",
             fontsize=12, fontweight="bold", color="#1A237E", va="top")
    info_line = (
        f"Type: {snap.get('sample_type', '-')}  |  "
        f"Species: {snap.get('species', '-')}  |  "
        f"Material: {snap.get('material', '-')}"
    )
    desc = (snap.get("description") or "")[:80]
    if desc:
        info_line += f"  |  {desc}"
    fig.text(0.048, 0.957, info_line, fontsize=7.5, color="#444444", va="top")
    if status:
        fig.text(0.952, 0.977, status,
                 fontsize=11, fontweight="bold", ha="right", va="top", color="white",
                 bbox=dict(facecolor=status_color, edgecolor="none",
                           boxstyle="round,pad=0.3"))

    # 헤더 구분선
    from matplotlib.lines import Line2D
    def _hline(y, color="#CCCCCC", lw=0.8, ls="-"):
        fig.add_artist(Line2D([0.048, 0.952], [y, y],
                              transform=fig.transFigure,
                              color=color, linewidth=lw, linestyle=ls))
    _hline(0.947)

    # ── QC Metrics 표 ─────────────────────────────────────────────────
    fig.text(0.048, 0.943, "QC Metrics",
             fontsize=9, fontweight="bold", color="#1A237E", va="top")

    QC_BOT = 0.710
    QC_H   = 0.228   # top ≈ 0.938
    ax_tbl = fig.add_axes([0.048, QC_BOT, 0.904, QC_H])
    ax_tbl.axis("off")

    tbl_col_labels = ["Step", "Instrument", "Conc\n(ng/µl)", "Vol\n(µl)",
                      "Total\n(ng)", "GQN/\nRIN", "Avg Size\n(bp)",
                      "%CV", "MQI", "Status", "Date"]

    smear_by_step: dict = {}
    for sa in smears:
        smear_by_step.setdefault(sa.step, {})[sa.range_text or ''] = sa

    if metrics_dicts:
        is_rna = 'rna' in (snap.get("sample_type") or '').lower()
        rows = []
        for m in metrics_dicts:
            date_str = m["measured_at"].strftime("%Y-%m-%d") if m["measured_at"] else "-"
            step_sm = smear_by_step.get(m["step"], {})
            cv_str = (
                _widest_cv(step_sm)
                if m.get("instrument") == "Femto Pulse" and is_rna else "-"
            )
            mqi_str = (
                _compute_mqi(step_sm)
                if m.get("instrument") == "Femto Pulse" and is_rna else "-"
            )
            rows.append([
                m["step"] or "-", m["instrument"] or "-",
                _fmt(m["concentration"]), _fmt(m["volume"]),
                _fmt(m["total_amount"]), _fmt(m["gqn_rin"]),
                _fmt(m["avg_size"], 0), cv_str, mqi_str,
                m["status"] or "-", date_str,
            ])
    else:
        rows = [["No data"] + ["-"] * (len(tbl_col_labels) - 1)]

    tbl = ax_tbl.table(
        cellText=rows, colLabels=tbl_col_labels,
        loc="upper center", cellLoc="center",
        bbox=[0.0, 0.0, 1.0, 1.0],
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(7)
    tbl.auto_set_column_width(list(range(len(tbl_col_labels))))
    _apply_table_style(
        tbl, len(tbl_col_labels), len(rows),
        status_col=tbl_col_labels.index("Status"),
        metrics_dicts_ref=metrics_dicts if metrics_dicts else [],
    )

    # ── 판정 기준 주석 ─────────────────────────────────────────────────
    criteria_text = _criteria_note(snap.get("sample_type", ""))
    if criteria_text:
        fig.text(0.048, QC_BOT - 0.007, criteria_text,
                 fontsize=6.0, color="#777777", va="top", style="italic")

    # ── 전기영동 섹션 ─────────────────────────────────────────────────
    if has_electro:
        _hline(0.692, color="#CCCCCC", lw=0.8, ls="--")
        fig.text(0.048, 0.688, "Electropherogram",
                 fontsize=9, fontweight="bold", color="#1A237E", va="top")

        EP_BOT = 0.352
        EP_H   = 0.330   # top ≈ 0.682
        try:
            ephero_fig, _, _, _, _ = qc_visualizer.plot_electropherogram_overlay(
                sid, traces, calibration
            )
            if ephero_fig is not None:
                buf = io.BytesIO()
                ephero_fig.savefig(buf, format="png", dpi=140,
                                   bbox_inches="tight", facecolor="white")
                plt.close(ephero_fig)
                buf.seek(0)
                img = plt.imread(buf)
                ax_ep = fig.add_axes([0.048, EP_BOT, 0.904, EP_H])
                ax_ep.imshow(img, aspect="auto")
                ax_ep.axis("off")
        except Exception as e:
            logger.error(f"Electropherogram embed failed for {sid}: {e}")

    # ── Smear Analysis ─────────────────────────────────────────────────
    if has_smear:
        SM_LABEL_Y = 0.344 if has_electro else (QC_BOT - 0.040)
        SM_BOT     = 0.038
        SM_H       = SM_LABEL_Y - 0.008 - SM_BOT

        _hline(SM_LABEL_Y + 0.004, color="#CCCCCC", lw=0.8, ls="--")
        fig.text(0.048, SM_LABEL_Y, "Smear Analysis",
                 fontsize=9, fontweight="bold", color="#1A237E", va="top")

        smear_cols = ["Step", "Range", "% Total", "Avg Size (bp)", "%CV", "DQN"]
        smear_data = [
            [
                s.step or "-",
                s.range_text or "-",
                f"{s.pct_total:.1f}" if s.pct_total is not None else "-",
                f"{s.avg_size:.0f}" if s.avg_size is not None else "-",
                f"{s.cv:.1f}" if s.cv is not None else "-",
                f"{s.dqn:.2f}" if s.dqn is not None else "-",
            ]
            for s in smears
        ]
        ax_sm = fig.add_axes([0.048, SM_BOT, 0.904, SM_H])
        ax_sm.axis("off")
        sm_tbl = ax_sm.table(
            cellText=smear_data,
            colLabels=smear_cols,
            loc="upper center", cellLoc="center",
            bbox=[0.0, 0.0, 1.0, 1.0],
        )
        sm_tbl.auto_set_font_size(False)
        sm_tbl.set_fontsize(7)
        sm_tbl.auto_set_column_width(list(range(len(smear_cols))))
        _apply_table_style(sm_tbl, len(smear_cols), len(smear_data))

    fig.text(0.5, 0.016,
             f"NGS Sample QC LIMS  |  {datetime.now().strftime('%Y-%m-%d')}",
             ha="center", fontsize=7, color="#888888")
    return fig


def _build_cover_page(selected_ids: List[str], snap_map: dict) -> "plt.Figure":
    """PDF 첫 페이지: 표지."""
    fig, ax = plt.subplots(figsize=(8.27, 11.69))
    fig.patch.set_facecolor("white")
    ax.axis("off")

    # 타이틀 & 날짜
    ax.text(0.5, 0.91, "NGS Sample QC Report",
            transform=ax.transAxes, fontsize=28, fontweight="bold",
            ha="center", color="#1A237E")
    ax.text(0.5, 0.86,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            transform=ax.transAxes, fontsize=12, ha="center", color="#555555")

    # 구분선
    ax.plot([0.048, 0.952], [0.84, 0.84],
            transform=ax.transAxes, color="#1A237E", linewidth=1.5)

    # 배치 통계
    counts = {"Pass": 0, "Warning": 0, "Fail": 0, "No Data": 0}
    for sid in selected_ids:
        st = snap_map.get(sid, {}).get("latest_status", "No Data")
        counts[st] = counts.get(st, 0) + 1

    ax.text(0.5, 0.80, f"Total Samples: {len(selected_ids)}",
            transform=ax.transAxes, fontsize=14, ha="center", fontweight="bold")

    stat_items = [
        ("Pass",    counts["Pass"]),
        ("Warning", counts["Warning"]),
        ("Fail",    counts["Fail"]),
        ("No Data", counts["No Data"]),
    ]
    for i, (st, cnt) in enumerate(stat_items):
        color = _STATUS_MPL.get(st, "#9E9E9E")
        ax.text(0.5, 0.75 - i * 0.048, f"{st}: {cnt}",
                transform=ax.transAxes, fontsize=12, ha="center",
                color=color, fontweight="bold")

    # 샘플 목록 테이블
    tbl_data = []
    for sid in selected_ids:
        s = snap_map.get(sid, {})
        tbl_data.append([
            sid,
            s.get("sample_name", "-"),
            s.get("sample_type", "-"),
            s.get("latest_status", "-"),
        ])

    # 표 높이를 행 수에 맞게 조정 (최대 0.40)
    tbl_height = min(0.40, 0.042 * len(tbl_data) + 0.06)
    tbl_bottom = 0.52 - tbl_height

    tbl = ax.table(
        cellText=tbl_data,
        colLabels=["Sample ID", "Name", "Type", "Status"],
        loc="center", cellLoc="center",
        bbox=[0.048, tbl_bottom, 0.904, tbl_height],
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8)
    tbl.auto_set_column_width([0, 1, 2, 3])

    for col in range(4):
        tbl[0, col].set_facecolor("#1A237E")
        tbl[0, col].set_text_props(color="white", fontweight="bold")

    for r, row_data in enumerate(tbl_data):
        status = row_data[3]
        color = _STATUS_MPL.get(status)
        if color:
            tbl[r + 1, 3].set_facecolor(color)
            tbl[r + 1, 3].set_text_props(color="white", fontweight="bold")
        if r % 2 == 1:
            for col in range(3):
                tbl[r + 1, col].set_facecolor("#F3F4F6")

    # 푸터
    ax.text(0.5, 0.014,
            "NGS Sample QC LIMS",
            transform=ax.transAxes, ha="center", fontsize=8, color="#888888")

    return fig


_STEP_ABBREV = {
    "gDNA Extraction":    "gDNA\nExt.",
    "SRE":                "SRE",
    "DNA Shearing":       "Shearing",
    "Library Prep":       "Lib.\nPrep",
    "Polymerase Binding": "Pol.\nBind.",
    "RNA Extraction":     "RNA\nExt.",
    "mRNA Elution":       "mRNA\nElut.",
    "Library Prep (RNA)": "Lib.\nPrep\n(RNA)",
}

# Standard step order (DNA first, then RNA)
_ALL_STEPS_ORDERED = list(QC_STEPS) + list(RNA_QC_STEPS)


def _build_summary_page(sample_ids: List[str], snap_map: dict) -> "plt.Figure":
    """PDF 두 번째 페이지: 배치 요약 (샘플 × QC step status). Portrait A4."""
    _order = {"Pass": 0, "Warning": 1, "Fail": 2}
    step_statuses: dict = {}   # {sid: {step: worst_status}}

    for sid in sample_ids:
        by_step: dict = {}
        try:
            with db_manager.session_scope() as session:
                metrics = get_qc_metrics_by_sample(session, sid)
                for m in metrics:
                    s = m.status
                    if s in _order:
                        prev = by_step.get(m.step)
                        if prev is None or _order[s] > _order.get(prev, -1):
                            by_step[m.step] = s
        except Exception:
            pass
        step_statuses[sid] = by_step

    # 실제 데이터가 있는 step만 표준 순서로 추출
    steps_with_data: set = set()
    for by_step in step_statuses.values():
        steps_with_data.update(by_step.keys())

    step_list = [s for s in _ALL_STEPS_ORDERED if s in steps_with_data]
    # 표준 목록에 없는 step은 끝에 추가
    for s in steps_with_data:
        if s not in step_list:
            step_list.append(s)

    step_labels = [_STEP_ABBREV.get(s, s) for s in step_list]

    # 정보 컬럼: Sample ID, Name, Type
    info_cols = ["Sample ID", "Name", "Type"]
    tbl_cols = info_cols + step_labels + ["Overall"]
    n_info = len(info_cols)

    tbl_data = []
    for sid in sample_ids:
        s = snap_map.get(sid, {})
        row = [sid, s.get("sample_name", "-"), s.get("sample_type", "-")]
        for step in step_list:
            row.append(step_statuses.get(sid, {}).get(step, "-"))
        row.append(s.get("latest_status", "-"))
        tbl_data.append(row)

    # Portrait A4
    fig, ax = plt.subplots(figsize=(8.27, 11.69))
    fig.patch.set_facecolor("white")
    ax.axis("off")

    ax.text(0.048, 0.966, "QC Batch Summary",
            transform=ax.transAxes, fontsize=14, fontweight="bold",
            color="#1A237E", va="top")
    ax.text(0.952, 0.966,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
            f"Total: {len(sample_ids)} sample(s)",
            transform=ax.transAxes, fontsize=8, ha="right", va="top", color="#555555")

    # 행 수에 맞게 표 높이 결정 (최소 헤더 포함)
    n_rows = len(tbl_data)
    tbl_height = min(0.888, 0.05 + n_rows * 0.055)
    tbl_bottom = 0.948 - tbl_height

    tbl = ax.table(
        cellText=tbl_data,
        colLabels=tbl_cols,
        loc="upper center",
        cellLoc="center",
        bbox=[0.048, tbl_bottom, 0.904, tbl_height],
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(7)
    tbl.auto_set_column_width(list(range(len(tbl_cols))))

    # 헤더 행 스타일
    for col in range(len(tbl_cols)):
        cell = tbl[0, col]
        cell.set_facecolor("#1A237E")
        cell.set_text_props(color="white", fontweight="bold", fontsize=7)

    status_col_indices = list(range(n_info, len(tbl_cols)))

    for row_idx, row_data in enumerate(tbl_data):
        for col_idx in status_col_indices:
            st = row_data[col_idx]
            color = _STATUS_MPL.get(st)
            if color and st != "-":
                tbl[row_idx + 1, col_idx].set_facecolor(color)
                tbl[row_idx + 1, col_idx].set_text_props(color="white", fontweight="bold")
        if row_idx % 2 == 1:
            for col in range(n_info):
                tbl[row_idx + 1, col].set_facecolor("#F3F4F6")

    ax.text(0.5, 0.014, "NGS Sample QC LIMS",
            transform=ax.transAxes, ha="center", fontsize=7, color="#888888")

    return fig


def _build_report_figure(snap: dict, metrics: list) -> "plt.Figure":
    """샘플 한 개의 QC 메트릭 페이지 Figure 생성."""
    fig = plt.figure(figsize=(8.27, 11.69))          # A4
    fig.patch.set_facecolor("white")

    gs = gridspec.GridSpec(
        4, 1, figure=fig,
        height_ratios=[1.2, 3, 2.8, 0.2],
        hspace=0.70,
        left=0.08, right=0.95, top=0.94, bottom=0.04,
    )

    # ── 1) 헤더 ─────────────────────────────────────────────────────
    ax_hdr = fig.add_subplot(gs[0])
    ax_hdr.axis("off")

    sample_id = snap.get("sample_id", "")
    status = snap.get("latest_status", "")
    status_color = _STATUS_MPL.get(status, "#9E9E9E")

    ax_hdr.text(0, 1.0, "NGS Sample QC Report",
                transform=ax_hdr.transAxes, fontsize=15, fontweight="bold",
                va="top", color="#1A237E")
    ax_hdr.text(0, 0.62, f"Sample ID: {sample_id}",
                transform=ax_hdr.transAxes, fontsize=11, va="top")

    info_lines = [
        f"Name: {snap.get('sample_name') or '-'}",
        f"Type: {snap.get('sample_type') or '-'}",
        f"Species: {snap.get('species') or '-'}",
        f"Material: {snap.get('material') or '-'}",
        f"Description: {snap.get('description') or '-'}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    ax_hdr.text(0, 0.30, "   ".join(info_lines[:3]),
                transform=ax_hdr.transAxes, fontsize=8, va="top", color="#333333")
    ax_hdr.text(0, 0.0, "   ".join(info_lines[3:]),
                transform=ax_hdr.transAxes, fontsize=8, va="top", color="#333333")

    # 상태 배지
    ax_hdr.text(1.0, 1.0, status,
                transform=ax_hdr.transAxes, fontsize=13, fontweight="bold",
                ha="right", va="top", color="white",
                bbox=dict(facecolor=status_color, edgecolor="none",
                          boxstyle="round,pad=0.4"))

    # ── 2) 차트 ─────────────────────────────────────────────────────
    ax_charts = fig.add_subplot(gs[1])
    ax_charts.axis("off")

    inner_gs = gridspec.GridSpecFromSubplotSpec(
        1, 3, subplot_spec=gs[1], wspace=0.52
    )
    ax_conc = fig.add_subplot(inner_gs[0])
    ax_gqn  = fig.add_subplot(inner_gs[1])
    ax_size = fig.add_subplot(inner_gs[2])

    if metrics:
        # Two-line x-tick labels: step name + instrument abbreviation
        steps = [
            f"{(m['step'] or '?')}\n({_short_instrument(m.get('instrument', ''))})"
            if m.get('instrument') else (m['step'] or '?')
            for m in metrics
        ]
        concs      = [m["concentration"] or 0 for m in metrics]
        gqns       = [m["gqn_rin"] or 0 for m in metrics]
        sizes      = [m["avg_size"] or 0 for m in metrics]
        statuses   = [m["status"] or "No Data" for m in metrics]
        bar_colors = [_STATUS_MPL.get(s, "#9E9E9E") for s in statuses]
        x = np.arange(len(steps))

        for ax, vals, ylabel in [
            (ax_conc, concs, "ng/µl"),
            (ax_gqn,  gqns,  "GQN/RIN"),
            (ax_size, sizes, "bp"),
        ]:
            bars = ax.bar(x, vals, color=bar_colors, edgecolor="white", linewidth=0.6)
            ax.set_xticks(x)
            ax.set_xticklabels(steps, rotation=35, ha="right", fontsize=6,
                               linespacing=1.2)
            ax.set_ylabel(ylabel, fontsize=7)
            ax.tick_params(axis="y", labelsize=7)
            ax.spines[["top", "right"]].set_visible(False)
            ax.set_facecolor("#FAFAFA")
            # 25% headroom so value annotations don't overflow into chart title
            max_v = max((v for v in vals if v), default=0)
            ax.set_ylim(0, max_v * 1.28 if max_v > 0 else 1)
            for bar, val in zip(bars, vals):
                if val > 0:
                    ax.text(bar.get_x() + bar.get_width() / 2,
                            bar.get_height(), f"{val:.1f}",
                            ha="center", va="bottom", fontsize=6)

        ax_conc.set_title("Concentration", fontsize=8, fontweight="bold")
        ax_gqn.set_title("GQN / RIN", fontsize=8, fontweight="bold")
        ax_size.set_title("Avg Size (bp)", fontsize=8, fontweight="bold")

        # 범례 (Pass/Warning/Fail)
        from matplotlib.patches import Patch
        legend_handles = [
            Patch(facecolor=_STATUS_MPL["Pass"],    label="Pass"),
            Patch(facecolor=_STATUS_MPL["Warning"], label="Warning"),
            Patch(facecolor=_STATUS_MPL["Fail"],    label="Fail"),
            Patch(facecolor=_STATUS_MPL["No Data"], label="No Data"),
        ]
        ax_conc.legend(handles=legend_handles, loc="upper right",
                       fontsize=6, frameon=True, framealpha=0.8)
    else:
        for ax in (ax_conc, ax_gqn, ax_size):
            ax.text(0.5, 0.5, "No QC data", ha="center", va="center",
                    transform=ax.transAxes, color="#9E9E9E", fontsize=9)
            ax.axis("off")

    # ── 3) QC 데이터 테이블 ─────────────────────────────────────────
    ax_tbl = fig.add_subplot(gs[2])
    ax_tbl.axis("off")

    tbl_cols = ["Step", "Instrument", "Conc\n(ng/µl)", "Vol\n(µl)",
                "Total\n(ng)", "GQN/\nRIN", "Avg Size\n(bp)",
                "Status", "Date"]

    if metrics:
        tbl_data = []
        for m in metrics:
            date_str = (m["measured_at"].strftime("%Y-%m-%d")
                        if m["measured_at"] else "-")
            tbl_data.append([
                m["step"] or "-",
                m["instrument"] or "-",
                _fmt(m["concentration"]),
                _fmt(m["volume"]),
                _fmt(m["total_amount"]),
                _fmt(m["gqn_rin"]),
                _fmt(m["avg_size"], 0),
                m["status"] or "-",
                date_str,
            ])
    else:
        tbl_data = [["No data"] + ["-"] * (len(tbl_cols) - 1)]

    tbl = ax_tbl.table(
        cellText=tbl_data,
        colLabels=tbl_cols,
        loc="upper center",
        cellLoc="center",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(7)
    tbl.auto_set_column_width(list(range(len(tbl_cols))))

    # 헤더 스타일
    for col in range(len(tbl_cols)):
        cell = tbl[0, col]
        cell.set_facecolor("#1A237E")
        cell.set_text_props(color="white", fontweight="bold")

    # 상태 셀 색상
    for row_idx, m in enumerate(metrics if metrics else []):
        status = m.get("status", "")
        color = _STATUS_MPL.get(status)
        if color:
            status_col = tbl_cols.index("Status")
            tbl[row_idx + 1, status_col].set_facecolor(color)
            tbl[row_idx + 1, status_col].set_text_props(
                color="white", fontweight="bold"
            )

    # 교대 행 배경
    for row_idx in range(len(metrics if metrics else [])):
        if row_idx % 2 == 1:
            for col in range(len(tbl_cols)):
                if (row_idx + 1, col) in tbl._cells:
                    cell = tbl[row_idx + 1, col]
                    if cell.get_facecolor()[:3] == (1, 1, 1):  # white만
                        cell.set_facecolor("#F3F4F6")

    ax_tbl.set_title("QC Metrics", fontsize=9, fontweight="bold",
                     loc="left", pad=4, color="#1A237E")

    # ── 4) 푸터 ─────────────────────────────────────────────────────
    ax_ftr = fig.add_subplot(gs[3])
    ax_ftr.axis("off")
    ax_ftr.text(0.5, 0.5,
                f"NGS Sample QC LIMS  |  {datetime.now().strftime('%Y-%m-%d')}",
                transform=ax_ftr.transAxes,
                ha="center", va="center", fontsize=7, color="#888888")

    return fig


def _build_sample_electro_page(
    sample_id: str,
    snap: dict,
    session,
) -> "Optional[plt.Figure]":
    """Electropherogram + Smear Analysis 페이지. 데이터 없으면 None 반환."""
    try:
        from analysis.visualizer import load_electropherogram_traces, qc_visualizer
    except ImportError:
        logger.warning("visualizer import failed — skipping electropherogram page")
        return None

    try:
        traces, calibration = load_electropherogram_traces(sample_id)
    except Exception as e:
        logger.error(f"load_electropherogram_traces failed for {sample_id}: {e}")
        return None

    if not traces:
        return None

    fig, ax, lines_dict, cal_bps, cal_times = qc_visualizer.plot_electropherogram_overlay(
        sample_id, traces, calibration
    )
    if fig is None:
        return None

    # A4 세로 크기로 조정
    fig.set_size_inches(8.27, 11.69)
    fig.patch.set_facecolor("white")

    # 헤더 텍스트 (fig 좌표계)
    status = snap.get("latest_status", "")
    status_color = _STATUS_MPL.get(status, "#9E9E9E")

    fig.text(0.06, 0.975, f"Electropherogram — {sample_id}",
             fontsize=13, fontweight="bold", color="#1A237E", va="top")
    fig.text(0.06, 0.955,
             f"{snap.get('sample_name', '')}  |  {snap.get('sample_type', '')}  |  "
             f"{snap.get('species', '')}",
             fontsize=9, color="#333333", va="top")
    if status:
        fig.text(0.94, 0.975, status,
                 fontsize=11, fontweight="bold", ha="right", va="top", color="white",
                 bbox=dict(facecolor=status_color, edgecolor="none",
                           boxstyle="round,pad=0.3"))

    # 차트 영역을 위로 올려 smear table 공간 확보
    ax.set_position([0.08, 0.36, 0.88, 0.55])

    # Smear Analysis 테이블
    try:
        smears = get_smear_analyses_by_sample(session, sample_id)
    except Exception as e:
        logger.error(f"get_smear_analyses_by_sample failed for {sample_id}: {e}")
        smears = []

    if smears:
        smear_data = [
            [
                s.step or "-",
                s.range_text or "-",
                f"{s.pct_total:.1f}" if s.pct_total is not None else "-",
                f"{s.avg_size:.0f}" if s.avg_size is not None else "-",
                f"{s.dqn:.2f}" if s.dqn is not None else "-",
            ]
            for s in smears
        ]
        smear_ax = fig.add_axes([0.06, 0.05, 0.88, 0.27])
        smear_ax.axis("off")
        smear_ax.set_title("Smear Analysis", fontsize=9, fontweight="bold",
                           loc="left", color="#1A237E", pad=4)

        smear_tbl = smear_ax.table(
            cellText=smear_data,
            colLabels=["Step", "Range", "% of Total", "Avg Size (bp)", "DQN"],
            loc="upper center",
            cellLoc="center",
        )
        smear_tbl.auto_set_font_size(False)
        smear_tbl.set_fontsize(7.5)
        smear_tbl.auto_set_column_width([0, 1, 2, 3, 4])

        for col in range(5):
            smear_tbl[0, col].set_facecolor("#1A237E")
            smear_tbl[0, col].set_text_props(color="white", fontweight="bold")

        for r in range(len(smear_data)):
            if r % 2 == 1:
                for col in range(5):
                    smear_tbl[r + 1, col].set_facecolor("#F3F4F6")

    # 푸터
    fig.text(0.5, 0.01,
             f"NGS Sample QC LIMS  |  {datetime.now().strftime('%Y-%m-%d')}",
             ha="center", fontsize=7, color="#888888")

    return fig


# ════════════════════════════════════════════════════════════════════
# Excel 빌더 함수
# ════════════════════════════════════════════════════════════════════

def _write_excel(wb: "openpyxl.Workbook", sample_ids: List[str], snap_map: dict):
    """선택 샘플의 QC 데이터를 Excel에 기록."""
    # ── Sheet 1: Sample Info ──────────────────────────────────────
    ws_info = wb.active
    ws_info.title = "Sample Info"

    hdr_fill = PatternFill("solid", fgColor="1A237E")
    hdr_font = XlFont(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="F3F4F6")
    center   = Alignment(horizontal="center", vertical="center")
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, row, cols):
        for col_idx, text in enumerate(cols, 1):
            c = ws.cell(row=row, column=col_idx, value=text)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = border

    def _cell(ws, row, col, value, fill=None):
        c = ws.cell(row=row, column=col, value=value)
        c.alignment = Alignment(vertical="center")
        c.border = border
        if fill:
            c.fill = fill
        return c

    info_cols = ["Sample ID", "Name", "Type", "Species", "Material",
                 "Description", "Latest Status"]
    _hdr(ws_info, 1, info_cols)

    for row_idx, sid in enumerate(sample_ids, 2):
        s = snap_map.get(sid, {})
        status = s.get("latest_status", "")
        status_color_map = {
            "Pass": "4CAF50", "Warning": "FF9800",
            "Fail": "F44336", "No Data": "9E9E9E"
        }
        status_fill = PatternFill("solid",
                                   fgColor=status_color_map.get(status, "9E9E9E"))
        alt = alt_fill if row_idx % 2 == 0 else None
        _cell(ws_info, row_idx, 1, sid, alt)
        _cell(ws_info, row_idx, 2, s.get("sample_name", ""), alt)
        _cell(ws_info, row_idx, 3, s.get("sample_type", ""), alt)
        _cell(ws_info, row_idx, 4, s.get("species", ""), alt)
        _cell(ws_info, row_idx, 5, s.get("material", ""), alt)
        _cell(ws_info, row_idx, 6, s.get("description", ""), alt)
        c = _cell(ws_info, row_idx, 7, status, status_fill)
        c.font = XlFont(bold=True, color="FFFFFF")
        c.alignment = center

    for col in ws_info.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws_info.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Sheet 2: QC Metrics ───────────────────────────────────────
    ws_qc = wb.create_sheet("QC Metrics")
    qc_cols = ["Sample ID", "Step", "Instrument",
               "Conc (ng/ul)", "Vol (ul)", "Total (ng)",
               "260/280", "GQN/RIN", "Avg Size (bp)",
               "Status", "Measured Date"]
    _hdr(ws_qc, 1, qc_cols)

    qc_row = 2
    status_color_map = {
        "Pass": "4CAF50", "Warning": "FF9800",
        "Fail": "F44336", "No Data": "9E9E9E"
    }
    for sid in sample_ids:
        try:
            with db_manager.session_scope() as session:
                metrics = get_qc_metrics_by_sample(session, sid)
                for m_idx, m in enumerate(metrics):
                    alt = alt_fill if qc_row % 2 == 0 else None
                    _cell(ws_qc, qc_row, 1, sid, alt)
                    _cell(ws_qc, qc_row, 2, m.step, alt)
                    _cell(ws_qc, qc_row, 3, m.instrument, alt)
                    _cell(ws_qc, qc_row, 4, m.concentration, alt)
                    _cell(ws_qc, qc_row, 5, m.volume, alt)
                    _cell(ws_qc, qc_row, 6, m.total_amount, alt)
                    _cell(ws_qc, qc_row, 7, m.purity_260_280, alt)
                    _cell(ws_qc, qc_row, 8, m.gqn_rin, alt)
                    _cell(ws_qc, qc_row, 9, m.avg_size, alt)
                    status = m.status or ""
                    s_fill = PatternFill("solid",
                                         fgColor=status_color_map.get(status, "9E9E9E"))
                    c = _cell(ws_qc, qc_row, 10, status, s_fill)
                    c.font = XlFont(bold=True, color="FFFFFF")
                    c.alignment = center
                    date_str = (m.measured_at.strftime("%Y-%m-%d")
                                if m.measured_at else "")
                    _cell(ws_qc, qc_row, 11, date_str, alt)
                    qc_row += 1
        except Exception as e:
            logger.error(f"Excel QC row failed for {sid}: {e}")

    for col in ws_qc.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws_qc.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    ws_qc.freeze_panes = "A2"
    ws_info.freeze_panes = "A2"
